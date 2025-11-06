#!/usr/bin/env python3
"""
JIRA Worklog Riport Készítő
Lekérdezi egy adott felhasználó worklogjait JIRA-ból és riportot készít
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import json
import os
from datetime import datetime
from typing import Dict, List, Optional
from collections import defaultdict
from jira import JIRA
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class JiraWorklogApp:
    def __init__(self, root):
        self.root = root
        self.root.title("JIRA Worklog Riport Készítő")
        self.root.geometry("800x600")
        
        # Adatok
        self.jira_config = None
        self.jira_client = None
        
        # GUI felépítése
        self.setup_ui()
        
        # Auth.json betöltése
        self.load_auth_config()
    
    def setup_ui(self):
        """GUI felület létrehozása"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Cím
        title_label = ttk.Label(
            main_frame, 
            text="JIRA Worklog Riport Készítő",
            font=('Helvetica', 16, 'bold')
        )
        title_label.grid(row=0, column=0, columnspan=2, pady=10)
        
        # Felhasználónév
        ttk.Label(main_frame, text="JIRA Felhasználónév(ek):").grid(
            row=1, column=0, sticky=tk.W, pady=5
        )
        self.username_entry = ttk.Entry(main_frame, width=40)
        self.username_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5)
        self.username_entry.insert(0, "kasnyikl")
        
        # Segítő szöveg
        help_label = ttk.Label(
            main_frame, 
            text="(Több felhasználó esetén vesszővel elválasztva: kasnyikl, izbekiz)",
            font=('Helvetica', 9),
            foreground='gray'
        )
        help_label.grid(row=1, column=2, sticky=tk.W, padx=5)
        
        # JQL Query
        ttk.Label(main_frame, text="JQL Lekérdezés:").grid(
            row=3, column=0, sticky=tk.W, pady=5
        )
        self.jql_entry = ttk.Entry(main_frame, width=40)
        self.jql_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5)
        self.jql_entry.insert(0, "project = MYPROJECT")
        
        # Gomb
        self.query_button = ttk.Button(
            main_frame,
            text="Lekérdezés indítása",
            command=self.run_query
        )
        self.query_button.grid(row=4, column=0, columnspan=3, pady=20)
        
        # Státusz
        ttk.Label(main_frame, text="Státusz:").grid(
            row=5, column=0, sticky=tk.W, pady=5
        )
        self.status_text = scrolledtext.ScrolledText(
            main_frame,
            height=15,
            width=70,
            state='disabled'
        )
        self.status_text.grid(row=6, column=0, columnspan=3, pady=5)
        
        # Progress bar
        self.progress = ttk.Progressbar(
            main_frame,
            mode='indeterminate',
            length=400
        )
        self.progress.grid(row=7, column=0, columnspan=3, pady=10)
        
        # Grid konfigurálása
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(6, weight=1)
    
    def load_auth_config(self):
        """Auth.json betöltése"""
        try:
            auth_file = os.path.join(os.path.dirname(__file__), 'auth.json')
            if not os.path.exists(auth_file):
                messagebox.showerror(
                    "Hiba",
                    f"Nem található az auth.json fájl!\n{auth_file}"
                )
                self.log_status("HIBA: auth.json nem található!")
                return
            
            with open(auth_file, 'r', encoding='utf-8') as f:
                config = json.load(f)
                self.jira_config = config.get('jira')
            
            if not self.jira_config:
                messagebox.showerror(
                    "Hiba",
                    "Hibás auth.json formátum!"
                )
                self.log_status("HIBA: Hibás auth.json formátum!")
                return
            
            self.log_status(f"Auth config betöltve: {self.jira_config['url']}")
            
        except Exception as e:
            messagebox.showerror("Hiba", f"Auth.json betöltési hiba: {str(e)}")
            self.log_status(f"HIBA: {str(e)}")
    
    def log_status(self, message: str):
        """Státusz naplózása"""
        self.status_text.configure(state='normal')
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.status_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.status_text.see(tk.END)
        self.status_text.configure(state='disabled')
        self.root.update()
    
    def connect_jira(self) -> bool:
        """Csatlakozás JIRA-hoz"""
        try:
            self.log_status("Csatlakozás JIRA-hoz...")
            
            self.jira_client = JIRA(
                server=self.jira_config['url'],
                token_auth=self.jira_config['pat']
            )
            
            # Kapcsolat tesztelése
            user = self.jira_client.myself()
            self.log_status(f"Sikeres csatlakozás! Bejelentkezve mint: {user['displayName']}")
            return True
            
        except Exception as e:
            messagebox.showerror("Hiba", f"JIRA csatlakozási hiba: {str(e)}")
            self.log_status(f"HIBA: {str(e)}")
            return False
    
    def fetch_worklogs(self, username: str, jql: str) -> List[Dict]:
        """Worklogok lekérdezése"""
        worklogs = []
        
        try:
            self.log_status(f"JQL keresés: {jql}")
            
            # JIRA issue-k lekérdezése
            start_at = 0
            max_results = 50
            total_issues = None
            
            while total_issues is None or start_at < total_issues:
                issues = self.jira_client.search_issues(
                    jql,
                    startAt=start_at,
                    maxResults=max_results,
                    fields='summary,worklog,project,issuetype,status'
                )
                
                if total_issues is None:
                    total_issues = issues.total
                    self.log_status(f"Összesen {total_issues} jegy található")
                
                self.log_status(f"Feldolgozás: {start_at + 1}-{min(start_at + max_results, total_issues)} / {total_issues}")
                
                # Worklogok feldolgozása minden issue-ban
                for issue in issues:
                    issue_worklogs = self.jira_client.worklogs(issue.key)
                    
                    for worklog in issue_worklogs:
                        # Csak a megadott felhasználó worklogjait
                        if worklog.author.name == username:
                            worklogs.append({
                                'issue_key': issue.key,
                                'issue_summary': issue.fields.summary,
                                'project': issue.fields.project.key,
                                'issue_type': issue.fields.issuetype.name,
                                'status': issue.fields.status.name,
                                'author': worklog.author.displayName,
                                'started': worklog.started,
                                'time_spent': worklog.timeSpent,
                                'time_spent_seconds': worklog.timeSpentSeconds,
                                'comment': getattr(worklog, 'comment', '')
                            })
                
                start_at += max_results
            
            self.log_status(f"Összesen {len(worklogs)} worklog bejegyzés található {username} felhasználónak")
            return worklogs
            
        except Exception as e:
            messagebox.showerror("Hiba", f"Worklog lekérdezési hiba: {str(e)}")
            self.log_status(f"HIBA: {str(e)}")
            return []
    
    def group_worklogs_by_issue(self, worklogs: List[Dict]) -> Dict:
        """Worklogok csoportosítása jegy szerint"""
        grouped = defaultdict(lambda: {
            'issue_summary': '',
            'project': '',
            'issue_type': '',
            'status': '',
            'worklogs': []
        })
        
        for worklog in worklogs:
            key = worklog['issue_key']
            grouped[key]['issue_summary'] = worklog['issue_summary']
            grouped[key]['project'] = worklog['project']
            grouped[key]['issue_type'] = worklog['issue_type']
            grouped[key]['status'] = worklog['status']
            grouped[key]['worklogs'].append(worklog)
        
        return dict(grouped)
    
    def calculate_monthly_stats(self, worklogs: List[Dict]) -> Dict:
        """Havi statisztikák számítása"""
        monthly_stats = defaultdict(lambda: {
            'issues': set(),
            'total_seconds': 0,
            'worklogs_count': 0
        })
        
        for worklog in worklogs:
            # Dátum parsírozása (JIRA formátum: 2024-11-06T10:30:00.000+0100)
            started_str = worklog['started']
            started_date = datetime.strptime(started_str[:19], "%Y-%m-%dT%H:%M:%S")
            month_key = started_date.strftime("%Y-%m")
            
            monthly_stats[month_key]['issues'].add(worklog['issue_key'])
            monthly_stats[month_key]['total_seconds'] += worklog['time_spent_seconds']
            monthly_stats[month_key]['worklogs_count'] += 1
        
        return dict(monthly_stats)
    
    def seconds_to_dhm(self, seconds: int) -> tuple:
        """Másodpercek konvertálása nap/óra/perc formátumra (8 órás munkanappal)"""
        # JIRA 8 órás munkanapokkal számol (1 nap = 8 óra = 28800 sec)
        days = seconds // (8 * 3600)
        seconds %= (8 * 3600)
        hours = seconds // 3600
        seconds %= 3600
        minutes = seconds // 60
        
        return days, hours, minutes
    
    def seconds_to_hours(self, seconds: int) -> float:
        """Másodpercek konvertálása órákra (tizedesjegyek)"""
        return round(seconds / 3600, 2)
    
    def create_excel_report(self, all_user_worklogs: Dict[str, List[Dict]], usernames: List[str]):
        """Excel riport készítése több munkalappal, felhasználónként elkülönítve"""
        try:
            # Reports mappa létrehozása
            reports_dir = os.path.join(os.path.dirname(__file__), 'reports')
            os.makedirs(reports_dir, exist_ok=True)
            
            # Fájlnév generálása
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            users_str = "_".join(usernames) if len(usernames) <= 3 else f"{len(usernames)}_users"
            filename = f"worklog_{users_str}_{timestamp}.xlsx"
            filepath = os.path.join(reports_dir, filename)
            
            self.log_status(f"Excel riport készítése: {filename}")
            
            # Workbook létrehozása
            wb = openpyxl.Workbook()
            wb.remove(wb.active)  # Alapértelmezett lap törlése
            
            # Stílusok
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            stat_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            issue_header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            summary_header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Összesítő adatok gyűjtése
            total_stats = {
                'total_issues': set(),
                'total_worklogs': 0,
                'total_seconds': 0,
                'user_stats': {}
            }
            
            # Minden felhasználóhoz munkalapok létrehozása
            for username in usernames:
                worklogs = all_user_worklogs.get(username, [])
                
                if not worklogs:
                    self.log_status(f"Nincs worklog {username} felhasználónak, kihagyva...")
                    continue
                
                self.log_status(f"Munkalapok létrehozása {username} felhasználónak...")
                
                # Felhasználó statisztikák
                total_stats['user_stats'][username] = {
                    'issues': set(),
                    'worklogs': len(worklogs),
                    'seconds': sum(wl['time_spent_seconds'] for wl in worklogs)
                }
                for wl in worklogs:
                    total_stats['user_stats'][username]['issues'].add(wl['issue_key'])
                    total_stats['total_issues'].add(wl['issue_key'])
                
                total_stats['total_worklogs'] += len(worklogs)
                total_stats['total_seconds'] += sum(wl['time_spent_seconds'] for wl in worklogs)
                
                # Felhasználónév rövidítése munkalap névhez (max 31 karakter Excel limit)
                sheet_prefix = username[:20] if len(username) > 20 else username
                
                # 1. MUNKALAP: Jegyek és Worklogok (felhasználónként)
                ws_issues = wb.create_sheet(f"{sheet_prefix} - Jegyek")
                grouped_worklogs = self.group_worklogs_by_issue(worklogs)
                
                row = 1
                for issue_key in sorted(grouped_worklogs.keys()):
                    issue_data = grouped_worklogs[issue_key]
                    
                    # Jegy fejléc
                    ws_issues.merge_cells(f'A{row}:G{row}')
                    cell = ws_issues.cell(row=row, column=1, 
                                         value=f"{issue_key} - {issue_data['issue_summary']}")
                    cell.fill = issue_header_fill
                    cell.font = Font(color="FFFFFF", bold=True, size=12)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    
                    row += 1
                    
                    # Jegy részletek
                    ws_issues.cell(row=row, column=1, value="Projekt:")
                    ws_issues.cell(row=row, column=2, value=issue_data['project'])
                    ws_issues.cell(row=row, column=3, value="Típus:")
                    ws_issues.cell(row=row, column=4, value=issue_data['issue_type'])
                    ws_issues.cell(row=row, column=5, value="Státusz:")
                    ws_issues.cell(row=row, column=6, value=issue_data['status'])
                    
                    for col in range(1, 8):
                        ws_issues.cell(row=row, column=col).font = Font(bold=True)
                    
                    row += 1
                    
                    # Worklog táblázat fejléc
                    worklog_headers = ['Dátum', 'Időtartam', 'Órák', 'Komment']
                    for col_num, header in enumerate(worklog_headers, 1):
                        cell = ws_issues.cell(row=row, column=col_num, value=header)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment
                        cell.border = border
                    
                    row += 1
                    
                    # Worklogok
                    total_seconds = 0
                    for wl in issue_data['worklogs']:
                        started_date = datetime.strptime(wl['started'][:19], "%Y-%m-%dT%H:%M:%S")
                        
                        ws_issues.cell(row=row, column=1, value=started_date.strftime("%Y-%m-%d %H:%M"))
                        ws_issues.cell(row=row, column=2, value=wl['time_spent'])
                        ws_issues.cell(row=row, column=3, value=self.seconds_to_hours(wl['time_spent_seconds']))
                        ws_issues.cell(row=row, column=4, value=wl['comment'])
                        
                        for col in range(1, 5):
                            ws_issues.cell(row=row, column=col).border = border
                        
                        total_seconds += wl['time_spent_seconds']
                        row += 1
                    
                    # Összesítés
                    days, hours, minutes = self.seconds_to_dhm(total_seconds)
                    total_hours = self.seconds_to_hours(total_seconds)
                    
                    ws_issues.cell(row=row, column=1, value="ÖSSZESEN:")
                    ws_issues.cell(row=row, column=2, value=f"{days}n {hours}ó {minutes}p")
                    ws_issues.cell(row=row, column=3, value=total_hours)
                    
                    for col in range(1, 4):
                        ws_issues.cell(row=row, column=col).font = Font(bold=True)
                        ws_issues.cell(row=row, column=col).border = border
                    
                    row += 2  # Üres sor a következő jegy előtt
                
                # Oszlopszélességek
                ws_issues.column_dimensions['A'].width = 20
                ws_issues.column_dimensions['B'].width = 15
                ws_issues.column_dimensions['C'].width = 12
                ws_issues.column_dimensions['D'].width = 60
                ws_issues.column_dimensions['E'].width = 15
                ws_issues.column_dimensions['F'].width = 15
                ws_issues.column_dimensions['G'].width = 15
                
                # 2. MUNKALAP: Havi Statisztika (felhasználónként)
                ws_stats = wb.create_sheet(f"{sheet_prefix} - Havi stat")
                monthly_stats = self.calculate_monthly_stats(worklogs)
                
                # Fejléc
                stat_headers = ['Hónap', 'Jegyek száma', 'Worklogok száma', 
                               'Napok', 'Órák', 'Percek', 'Összesen (óra)']
                for col_num, header in enumerate(stat_headers, 1):
                    cell = ws_stats.cell(row=1, column=col_num, value=header)
                    cell.fill = stat_header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Adatok
                row = 2
                for month in sorted(monthly_stats.keys()):
                    stats = monthly_stats[month]
                    days, hours, minutes = self.seconds_to_dhm(stats['total_seconds'])
                    total_hours = self.seconds_to_hours(stats['total_seconds'])
                    
                    ws_stats.cell(row=row, column=1, value=month)
                    ws_stats.cell(row=row, column=2, value=len(stats['issues']))
                    ws_stats.cell(row=row, column=3, value=stats['worklogs_count'])
                    ws_stats.cell(row=row, column=4, value=days)
                    ws_stats.cell(row=row, column=5, value=hours)
                    ws_stats.cell(row=row, column=6, value=minutes)
                    ws_stats.cell(row=row, column=7, value=total_hours)
                    
                    for col in range(1, 8):
                        ws_stats.cell(row=row, column=col).border = border
                        ws_stats.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                    
                    row += 1
                
                # Oszlopszélességek
                for i in range(1, 8):
                    ws_stats.column_dimensions[get_column_letter(i)].width = 18
                
                # 3. MUNKALAP: Összes Worklog (Részletes lista - felhasználónként)
                ws_all = wb.create_sheet(f"{sheet_prefix} - Részletes")
                
                # Fejléc
                headers = [
                    'Jegy kulcs', 'Jegy címe', 'Projekt', 'Típus', 'Státusz',
                    'Felhasználó', 'Dátum', 'Időtartam', 'Órák', 'Megjegyzés'
                ]
                
                for col_num, header in enumerate(headers, 1):
                    cell = ws_all.cell(row=1, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Adatok
                for row_num, worklog in enumerate(worklogs, 2):
                    ws_all.cell(row=row_num, column=1, value=worklog['issue_key'])
                    ws_all.cell(row=row_num, column=2, value=worklog['issue_summary'])
                    ws_all.cell(row=row_num, column=3, value=worklog['project'])
                    ws_all.cell(row=row_num, column=4, value=worklog['issue_type'])
                    ws_all.cell(row=row_num, column=5, value=worklog['status'])
                    ws_all.cell(row=row_num, column=6, value=worklog['author'])
                    ws_all.cell(row=row_num, column=7, value=worklog['started'])
                    ws_all.cell(row=row_num, column=8, value=worklog['time_spent'])
                    ws_all.cell(row=row_num, column=9, value=self.seconds_to_hours(worklog['time_spent_seconds']))
                    ws_all.cell(row=row_num, column=10, value=worklog['comment'])
                    
                    for col in range(1, 11):
                        ws_all.cell(row=row_num, column=col).border = border
                
                # Oszlopszélességek
                column_widths = [15, 50, 15, 15, 15, 25, 20, 15, 12, 50]
                for i, width in enumerate(column_widths, 1):
                    ws_all.column_dimensions[get_column_letter(i)].width = width
            
            # ÖSSZESÍTŐ MUNKALAP (ha több felhasználó van)
            if len(usernames) > 1:
                self.log_status("Összesítő munkalap létrehozása...")
                ws_summary = wb.create_sheet("ÖSSZESÍTŐ", 0)  # Első helyre
                
                # Fejléc
                ws_summary.merge_cells('A1:E1')
                title_cell = ws_summary.cell(row=1, column=1, value="FELHASZNÁLÓK ÖSSZESÍTÉSE")
                title_cell.fill = summary_header_fill
                title_cell.font = Font(color="000000", bold=True, size=14)
                title_cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Táblázat fejléc
                summary_headers = ['Felhasználó', 'Jegyek száma', 'Worklogok száma', 'Napok', 'Órák', 'Percek', 'Összesen (óra)']
                for col_num, header in enumerate(summary_headers, 1):
                    cell = ws_summary.cell(row=3, column=col_num, value=header)
                    cell.fill = summary_header_fill
                    cell.font = Font(bold=True)
                    cell.alignment = header_alignment
                    cell.border = border
                
                # Felhasználók adatai
                row = 4
                for username in usernames:
                    if username in total_stats['user_stats']:
                        stats = total_stats['user_stats'][username]
                        days, hours, minutes = self.seconds_to_dhm(stats['seconds'])
                        total_hours = self.seconds_to_hours(stats['seconds'])
                        
                        ws_summary.cell(row=row, column=1, value=username)
                        ws_summary.cell(row=row, column=2, value=len(stats['issues']))
                        ws_summary.cell(row=row, column=3, value=stats['worklogs'])
                        ws_summary.cell(row=row, column=4, value=days)
                        ws_summary.cell(row=row, column=5, value=hours)
                        ws_summary.cell(row=row, column=6, value=minutes)
                        ws_summary.cell(row=row, column=7, value=total_hours)
                        
                        for col in range(1, 8):
                            ws_summary.cell(row=row, column=col).border = border
                            ws_summary.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                        
                        row += 1
                
                # Összesen sor
                total_days, total_hours_val, total_minutes = self.seconds_to_dhm(total_stats['total_seconds'])
                total_hours_decimal = self.seconds_to_hours(total_stats['total_seconds'])
                
                ws_summary.cell(row=row, column=1, value="ÖSSZESEN:")
                ws_summary.cell(row=row, column=2, value=len(total_stats['total_issues']))
                ws_summary.cell(row=row, column=3, value=total_stats['total_worklogs'])
                ws_summary.cell(row=row, column=4, value=total_days)
                ws_summary.cell(row=row, column=5, value=total_hours_val)
                ws_summary.cell(row=row, column=6, value=total_minutes)
                ws_summary.cell(row=row, column=7, value=total_hours_decimal)
                
                for col in range(1, 8):
                    ws_summary.cell(row=row, column=col).font = Font(bold=True, size=12)
                    ws_summary.cell(row=row, column=col).border = border
                    ws_summary.cell(row=row, column=col).alignment = Alignment(horizontal="center")
                    ws_summary.cell(row=row, column=col).fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                # Oszlopszélességek
                ws_summary.column_dimensions['A'].width = 25
                for i in range(2, 8):
                    ws_summary.column_dimensions[get_column_letter(i)].width = 18
            
            # Mentés
            wb.save(filepath)
            self.log_status(f"Riport sikeresen elkészült: {filepath}")
            
            # Statisztikák összefoglalása
            summary_text = f"Riport sikeresen elkészült!\n\n"
            summary_text += f"• {len(usernames)} felhasználó\n"
            summary_text += f"• {len(total_stats['total_issues'])} különböző jegy\n"
            summary_text += f"• {total_stats['total_worklogs']} worklog bejegyzés\n"
            summary_text += f"• {self.seconds_to_hours(total_stats['total_seconds'])} óra összesen\n\n"
            summary_text += f"Fájl: {filename}"
            
            messagebox.showinfo("Siker", summary_text)
            
        except Exception as e:
            messagebox.showerror("Hiba", f"Excel riport készítési hiba: {str(e)}")
            self.log_status(f"HIBA: {str(e)}")
    
    def run_query(self):
        """Lekérdezés futtatása"""
        # Mezők ellenőrzése
        usernames_input = self.username_entry.get().strip()
        jql = self.jql_entry.get().strip()
        
        if not usernames_input:
            messagebox.showwarning("Figyelmeztetés", "Add meg a felhasználónevet!")
            return
        
        if not jql:
            messagebox.showwarning("Figyelmeztetés", "Add meg a JQL lekérdezést!")
            return
        
        if not self.jira_config:
            messagebox.showerror("Hiba", "Auth config nincs betöltve!")
            return
        
        # Felhasználónevek feldolgozása (vesszővel elválasztva)
        usernames = [u.strip() for u in usernames_input.split(',') if u.strip()]
        
        if not usernames:
            messagebox.showwarning("Figyelmeztetés", "Add meg legalább egy felhasználónevet!")
            return
        
        self.log_status(f"Lekérdezés {len(usernames)} felhasználóra: {', '.join(usernames)}")
        
        # Gomb letiltása
        self.query_button.config(state='disabled')
        self.progress.start()
        
        try:
            # JIRA csatlakozás
            if not self.connect_jira():
                return
            
            # Worklogok lekérdezése minden felhasználóra
            all_user_worklogs = {}
            total_worklogs_count = 0
            
            for username in usernames:
                self.log_status(f"Worklogok lekérdezése: {username}")
                worklogs = self.fetch_worklogs(username, jql)
                all_user_worklogs[username] = worklogs
                total_worklogs_count += len(worklogs)
            
            if total_worklogs_count == 0:
                messagebox.showinfo(
                    "Információ",
                    f"Nem található worklog bejegyzés a megadott felhasználóknak ({', '.join(usernames)}) a JQL szerint."
                )
                return
            
            # Riport készítése
            self.create_excel_report(all_user_worklogs, usernames)
            
        finally:
            # Gomb engedélyezése
            self.query_button.config(state='normal')
            self.progress.stop()


def main():
    root = tk.Tk()
    app = JiraWorklogApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
